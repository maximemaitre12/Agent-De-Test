"""
Agent Testeur QA — Version sans IA (Playwright)
================================================
• Ouvre un vrai navigateur Chromium
• Inspecte le HTML pour trouver les agents, champs et boutons
• Lit les tests depuis un fichier Excel
• Exécute chaque test et note OK/KO dans l'Excel
• Aucune dépendance IA / clé API
"""

import sys
import os
import time
import glob
from datetime import datetime
from pathlib import Path

# ── Dépendances requises ──────────────────────────────────────
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("ERREUR : playwright non installé.")
    print("Lance : pip install playwright && playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERREUR : openpyxl non installé.")
    print("Lance : pip install openpyxl")
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────
SITE_URL      = "https://elio.onepoint.cloud/consulting-agents"
TIMEOUT_MS    = 90_000   # temps max d'attente d'une réponse (90s)
NAV_TIMEOUT   = 30_000   # timeout de navigation
SLOW_MO       = 150      # ms entre chaque action (visible par l'utilisateur)

# Couleurs Excel pour les résultats
FILL_OK  = PatternFill("solid", fgColor="C6EFCE")   # vert pâle
FILL_KO  = PatternFill("solid", fgColor="FFC7CE")   # rouge pâle
FONT_OK  = Font(bold=True, color="276221")
FONT_KO  = Font(bold=True, color="9C0006")

# ── Recherche automatique du fichier Excel ────────────────────
def find_excel_file() -> str | None:
    """Cherche un fichier .xlsx dans le dossier du script et le Bureau."""
    script_dir = Path(__file__).parent
    desktop    = Path.home() / "Desktop"
    onedrive   = Path.home() / "OneDrive"

    patterns = [
        str(script_dir / "*.xlsx"),
        str(script_dir / "*.xls"),
        str(desktop / "*.xlsx"),
        str(desktop / "*.xls"),
    ]
    # Cherche aussi dans les sous-dossiers OneDrive/Bureau
    for od in onedrive.glob("**/Bureau"):
        patterns.append(str(od / "*.xlsx"))

    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            # Prend le plus récemment modifié
            return max(matches, key=os.path.getmtime)
    return None


# ── Lecture de l'Excel ────────────────────────────────────────
class TestCase:
    def __init__(self, row_idx: int, test_id: str, description: str,
                 input_text: str, expected: str,
                 col_status: int, col_comment: int):
        self.row_idx    = row_idx      # numéro de ligne dans la feuille (1-based)
        self.test_id    = test_id
        self.description= description
        self.input_text = input_text
        self.expected   = expected
        self.col_status = col_status   # colonne où écrire OK/KO
        self.col_comment= col_comment  # colonne où écrire l'observation


def detect_columns(sheet) -> dict:
    """
    Lit la première ligne pour détecter les colonnes importantes.
    Retourne un dict {role: col_index} avec les rôles :
      input, expected, status, comment, id, description
    """
    KEYWORDS = {
        "id":          ["id", "n°", "num", "#", "test_id", "cas"],
        "description": ["description", "action", "scénario", "test", "etape", "étape"],
        "input":       ["input", "saisie", "prompt", "question", "texte", "message", "entrée", "entree"],
        "expected":    ["attendu", "expected", "résultat attendu", "résultat", "critère", "critere"],
        "status":      ["statut", "status", "ok/ko", "résultat", "result", "ok", "ko"],
        "comment":     ["commentaire", "comment", "observation", "notes", "détail", "detail"],
    }

    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    mapping = {}

    for col_idx, cell_val in enumerate(header_row, start=1):
        if cell_val is None:
            continue
        cell_str = str(cell_val).lower().strip()
        for role, keywords in KEYWORDS.items():
            if role not in mapping:
                for kw in keywords:
                    if kw in cell_str:
                        mapping[role] = col_idx
                        break

    # Valeurs par défaut si colonnes non trouvées
    defaults = {"id": 1, "description": 2, "input": 3,
                "expected": 4, "status": 5, "comment": 6}
    for role, default_col in defaults.items():
        if role not in mapping:
            mapping[role] = default_col

    return mapping


def load_tests_from_excel(path: str) -> dict[str, list[TestCase]]:
    """
    Lit toutes les feuilles de l'Excel.
    Retourne {nom_feuille: [TestCase, ...]}
    """
    wb = openpyxl.load_workbook(path)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = detect_columns(ws)
        tests = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Ignore les lignes vides
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            def get(col): return str(row[col - 1]).strip() if row[col - 1] is not None else ""

            test_id     = get(cols["id"])
            description = get(cols["description"])
            input_text  = get(cols["input"])
            expected    = get(cols["expected"])

            if not input_text and not description:
                continue

            tests.append(TestCase(
                row_idx    = row_idx,
                test_id    = test_id,
                description= description,
                input_text = input_text or description,  # fallback sur description
                expected   = expected,
                col_status = cols["status"],
                col_comment= cols["comment"],
            ))

        if tests:
            result[sheet_name] = tests
            print(f"  [Excel] Feuille '{sheet_name}' : {len(tests)} test(s)")

    wb.close()
    return result


def write_result(path: str, sheet_name: str, row_idx: int,
                 col_status: int, col_comment: int,
                 status: str, comment: str):
    """Ouvre l'Excel, écrit le résultat dans la bonne cellule, sauvegarde."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]

    cell_status  = ws.cell(row=row_idx, column=col_status)
    cell_comment = ws.cell(row=row_idx, column=col_comment)

    cell_status.value = status
    cell_comment.value = comment

    if status == "OK":
        cell_status.fill = FILL_OK
        cell_status.font = FONT_OK
    else:
        cell_status.fill = FILL_KO
        cell_status.font = FONT_KO

    wb.save(path)
    wb.close()


# ── Auto-découverte de la page ────────────────────────────────

def find_agent_card(page, agent_name: str):
    """
    Cherche une carte/bouton dont le texte correspond à agent_name.
    Essaie plusieurs stratégies de sélection.
    """
    name_lower = agent_name.lower().strip()

    # Stratégie 1 : cherche un élément cliquable contenant exactement le texte
    for selector in [
        f"text='{agent_name}'",
        f"text={agent_name}",
        f"[aria-label*='{agent_name}' i]",
        f"h2:has-text('{agent_name}')",
        f"h3:has-text('{agent_name}')",
        f".card:has-text('{agent_name}')",
        f"button:has-text('{agent_name}')",
        f"a:has-text('{agent_name}')",
    ]:
        try:
            el = page.locator(selector).first
            if el.count() > 0 and el.is_visible(timeout=1000):
                return el
        except Exception:
            pass

    # Stratégie 2 : recherche floue dans tous les éléments cliquables
    for tag in ["button", "a", "div[role='button']", "[onclick]", ".card", "article"]:
        try:
            elements = page.locator(tag).all()
            for el in elements:
                try:
                    text = el.inner_text(timeout=500).lower()
                    if name_lower in text:
                        return el
                except Exception:
                    pass
        except Exception:
            pass

    return None


def find_input_field(page):
    """
    Détecte automatiquement le champ de saisie de l'agent.
    Priorité : textarea > input[type=text] > [contenteditable] > input visible
    """
    candidates = [
        "textarea:visible",
        "input[type='text']:visible",
        "input[type='search']:visible",
        "[contenteditable='true']:visible",
        "input:not([type='hidden']):not([type='submit']):not([type='button']):visible",
    ]
    for selector in candidates:
        try:
            el = page.locator(selector).first
            if el.count() > 0 and el.is_visible(timeout=1000):
                return el
        except Exception:
            pass
    return None


def find_submit_button(page, input_el=None):
    """
    Cherche le bouton d'envoi le plus proche du champ de saisie.
    """
    candidates = [
        "button[type='submit']:visible",
        "button:has-text('Envoyer'):visible",
        "button:has-text('Send'):visible",
        "button:has-text('Valider'):visible",
        "button:has-text('Go'):visible",
        "[aria-label*='send' i]:visible",
        "[aria-label*='envoyer' i]:visible",
        "[aria-label*='submit' i]:visible",
    ]
    for selector in candidates:
        try:
            el = page.locator(selector).first
            if el.count() > 0 and el.is_visible(timeout=1000):
                return el
        except Exception:
            pass
    return None


def wait_for_response_stable(page, timeout_ms: int = TIMEOUT_MS) -> str:
    """
    Attend que la page soit stable (réseau calme + plus de spinner).
    Retourne le texte du contenu principal de la page.
    """
    # Attend la fin des requêtes réseau
    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
    except PWTimeout:
        pass  # timeout acceptable, on continue

    # Attend la disparition des spinners classiques
    spinner_selectors = [
        ".spinner", ".loading", ".loader",
        "[aria-label*='loading' i]", "[aria-busy='true']",
        ".dots", ".typing-indicator", ".thinking",
    ]
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        visible_spinners = False
        for sel in spinner_selectors:
            try:
                if page.locator(sel).first.is_visible(timeout=200):
                    visible_spinners = True
                    break
            except Exception:
                pass
        if not visible_spinners:
            break
        time.sleep(1)

    # Attend que le texte se stabilise (plus de modifications pendant 2s)
    prev_text = ""
    stable_count = 0
    while time.time() < deadline:
        try:
            current_text = page.locator("body").inner_text(timeout=2000)
        except Exception:
            current_text = ""
        if current_text == prev_text:
            stable_count += 1
            if stable_count >= 2:
                break
        else:
            stable_count = 0
            prev_text = current_text
        time.sleep(1)

    return prev_text


def evaluate_result(response_text: str, expected: str, error_before: str) -> tuple[str, str]:
    """
    Détermine OK ou KO en comparant la réponse avec le résultat attendu.
    Retourne (status, comment).
    """
    response_lower = response_text.lower()
    expected_lower = expected.lower().strip()

    # Détection d'erreurs techniques
    error_keywords = ["erreur", "error", "500", "404", "exception",
                      "crash", "unavailable", "indisponible", "failed"]
    has_error = any(kw in response_lower for kw in error_keywords)

    # Si pas de critère attendu : vérifie juste qu'il y a une réponse et pas d'erreur
    if not expected_lower:
        if has_error:
            return "KO", "Erreur technique détectée dans la réponse."
        new_content = response_text.replace(error_before, "").strip()
        if len(new_content) > 20:
            return "OK", f"Réponse reçue ({len(new_content)} caractères). Aucun critère défini."
        return "KO", "Aucune réponse détectée."

    # Vérifie que les mots-clés attendus sont présents
    keywords = [kw.strip() for kw in expected_lower.replace(",", " ").split() if len(kw.strip()) > 3]
    matched = [kw for kw in keywords if kw in response_lower]

    if has_error:
        return "KO", f"Erreur technique détectée. Attendu : '{expected}'"

    if not keywords or len(matched) >= max(1, len(keywords) // 2):
        return "OK", f"Critères satisfaits ({len(matched)}/{len(keywords)} mots-clés trouvés)."

    return "KO", f"Critères non satisfaits. Attendu : '{expected}'. Trouvé : {len(matched)}/{len(keywords)} mots-clés."


def find_reset_button(page):
    """Cherche un bouton de réinitialisation/nouveau chat."""
    candidates = [
        "button:has-text('Nouveau'):visible",
        "button:has-text('Reset'):visible",
        "button:has-text('Effacer'):visible",
        "button:has-text('Clear'):visible",
        "button:has-text('New chat'):visible",
        "button:has-text('Nouvelle conversation'):visible",
        "[aria-label*='reset' i]:visible",
        "[aria-label*='clear' i]:visible",
        "[aria-label*='nouveau' i]:visible",
        "[title*='reset' i]:visible",
        "[title*='effacer' i]:visible",
    ]
    for selector in candidates:
        try:
            el = page.locator(selector).first
            if el.count() > 0 and el.is_visible(timeout=500):
                return el
        except Exception:
            pass
    return None


# ── Boucle principale ─────────────────────────────────────────

def run():
    print("═" * 62)
    print("  AGENT TESTEUR QA — Sans IA (Playwright)")
    print("═" * 62)
    print(f"  Démarrage : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print()

    # 1. Trouver le fichier Excel
    excel_path = find_excel_file()
    if not excel_path:
        excel_path = input("  Fichier Excel non trouvé. Entre le chemin complet : ").strip().strip('"')
    if not os.path.exists(excel_path):
        print(f"  ERREUR : fichier introuvable → {excel_path}")
        sys.exit(1)
    print(f"  Excel trouvé : {excel_path}")
    print()

    # 2. Charger les tests
    all_tests = load_tests_from_excel(excel_path)
    if not all_tests:
        print("  ERREUR : aucun test trouvé dans l'Excel.")
        sys.exit(1)
    print(f"  {sum(len(v) for v in all_tests.values())} test(s) au total dans {len(all_tests)} agent(s).")
    print()
    input("  Appuie sur ENTRÉE pour lancer le navigateur...")
    print()

    # 3. Lancer Playwright
    total_ok  = 0
    total_ko  = 0
    total_all = 0

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, slow_mo=SLOW_MO)
        context = browser.new_context(viewport={"width": 1280, "height": 800})
        page    = context.new_page()
        page.set_default_timeout(NAV_TIMEOUT)

        print(f"  Navigation vers {SITE_URL}...")
        page.goto(SITE_URL)
        page.wait_for_load_state("networkidle")
        print("  Page d'accueil chargée.")
        print()

        # 4. Pour chaque agent (feuille Excel)
        for agent_name, tests in all_tests.items():
            print(f"┌─ Agent : '{agent_name}' ({len(tests)} test(s))")

            # Revenir à la page d'accueil
            if page.url != SITE_URL:
                page.goto(SITE_URL)
                page.wait_for_load_state("networkidle")

            # Trouver la carte de l'agent
            card = find_agent_card(page, agent_name)
            if card is None:
                print(f"│  ⚠️  Agent '{agent_name}' non trouvé sur la page. Passage au suivant.")
                for tc in tests:
                    write_result(excel_path, agent_name, tc.row_idx,
                                 tc.col_status, tc.col_comment,
                                 "KO", f"Agent '{agent_name}' introuvable sur {SITE_URL}")
                    total_ko  += 1
                    total_all += 1
                print(f"└─ {len(tests)} test(s) marqués KO.")
                continue

            # Ouvrir l'agent
            try:
                card.click()
                page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT)
                time.sleep(1)
            except Exception as e:
                print(f"│  ⚠️  Impossible d'ouvrir l'agent : {e}")
                continue

            print(f"│  Agent ouvert — URL : {page.url}")

            agent_ok = 0
            agent_ko = 0

            # 5. Pour chaque test
            for tc in tests:
                label = f"[{tc.test_id}]" if tc.test_id else f"[ligne {tc.row_idx}]"
                print(f"│  ├─ Test {label} : {tc.description[:60]}")

                # Reset de l'agent avant chaque test
                reset_btn = find_reset_button(page)
                if reset_btn:
                    try:
                        reset_btn.click()
                        time.sleep(0.8)
                    except Exception:
                        pass
                else:
                    # Fallback : recharger la page et rouvrir l'agent
                    page.goto(SITE_URL)
                    page.wait_for_load_state("networkidle")
                    card = find_agent_card(page, agent_name)
                    if card:
                        card.click()
                        page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT)
                        time.sleep(1)

                # Capturer le texte AVANT la saisie (pour diff)
                try:
                    text_before = page.locator("body").inner_text(timeout=3000)
                except Exception:
                    text_before = ""

                # Trouver le champ de saisie
                input_el = find_input_field(page)
                if input_el is None:
                    status  = "KO"
                    comment = "Champ de saisie introuvable sur la page."
                    print(f"│  │    → {status} — {comment}")
                    write_result(excel_path, agent_name, tc.row_idx,
                                 tc.col_status, tc.col_comment, status, comment)
                    agent_ko  += 1
                    total_all += 1
                    total_ko  += 1
                    continue

                # Saisir le texte
                try:
                    input_el.click()
                    input_el.fill(tc.input_text)
                except Exception as e:
                    status  = "KO"
                    comment = f"Impossible de saisir le texte : {e}"
                    print(f"│  │    → {status} — {comment}")
                    write_result(excel_path, agent_name, tc.row_idx,
                                 tc.col_status, tc.col_comment, status, comment)
                    agent_ko  += 1
                    total_all += 1
                    total_ko  += 1
                    continue

                # Envoyer : bouton submit ou touche Entrée
                submit_btn = find_submit_button(page, input_el)
                try:
                    if submit_btn:
                        submit_btn.click()
                    else:
                        input_el.press("Enter")
                except Exception:
                    pass  # on continue même si le clic échoue

                # Attendre la réponse
                print(f"│  │    En attente de réponse (max {TIMEOUT_MS//1000}s)...")
                response_text = wait_for_response_stable(page, TIMEOUT_MS)

                # Évaluer le résultat
                status, comment = evaluate_result(response_text, tc.expected, text_before)

                icon = "✅" if status == "OK" else "❌"
                print(f"│  │    → {icon} {status} — {comment}")

                # Écrire dans l'Excel
                write_result(excel_path, agent_name, tc.row_idx,
                             tc.col_status, tc.col_comment, status, comment)

                if status == "OK":
                    agent_ok  += 1
                    total_ok  += 1
                else:
                    agent_ko  += 1
                    total_ko  += 1
                total_all += 1

            print(f"└─ Bilan '{agent_name}' : {agent_ok} OK / {agent_ko} KO")
            print()

        browser.close()

    # 6. Rapport final
    print("═" * 62)
    print("  RAPPORT FINAL")
    print("═" * 62)
    print(f"  Tests exécutés : {total_all}")
    print(f"  OK             : {total_ok}")
    print(f"  KO             : {total_ko}")
    if total_all > 0:
        print(f"  Taux de succès : {round(total_ok / total_all * 100)}%")
    print(f"  Résultats sauvegardés dans : {excel_path}")
    print(f"  Fin : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("═" * 62)


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\n\n⛔ Arrêt manuel (Ctrl+C).")
        sys.exit(0)
